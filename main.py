import asyncio
import os
import re
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, F, Router
from aiogram.types import Message, CallbackQuery, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove, FSInputFile
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.client.default import DefaultBotProperties
from openpyxl import Workbook, load_workbook

# BAZA ULANISHI
try:
    from db_manager import db
except ImportError:
    print("XATO: db_manager.py fayli topilmadi!")
    exit()

# --- SOZLAMALAR ---
TOKEN = "8143822107:AAFgSsJMeJ9SGdf1dQflBnExlvnsBIfRdzs"
ADMIN_IDS = [7044905076, 6134534264]
ADMIN_PASSWORD = "1122"
COADMIN_PASSWORD = "3344"

# --- STATES ---
class Reg(StatesGroup): name = State()
class AdminSt(StatesGroup): password = State(); add_nums = State(); set_limit = State()
class CoAdminSt(StatesGroup): password = State()
class Call(StatesGroup): 
    waiting_for_action = State(); phone = State(); status = State(); name = State()
    age = State(); height = State(); weight = State(); interest = State()
class ManualEntry(StatesGroup): phone = State(); action = State()
class Search(StatesGroup): query = State()

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
dp = Dispatcher(storage=MemoryStorage())
router = Router()
dp.include_router(router)

# ================== KLAVIATURALAR ==================
back_kb = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="🔙 Ortga")]], resize_keyboard=True)

def main_kb(uid, online=False):
    status_text = "Online ✅" if online else "Offline ❌"
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📞 Nomer olish"), KeyboardButton(text="✍️ Raqam kiritish")],
        [KeyboardButton(text="🔍 Raqam tekshirish")],
        [KeyboardButton(text="🟢 Ish vaqti: " + status_text)],
        [KeyboardButton(text="📉 Bugungi qabul qilinmagan")],
        [KeyboardButton(text="📈 Shaxsiy statistikam")]
    ], resize_keyboard=True)

call_action_kb = ReplyKeyboardMarkup(keyboard=[
    [KeyboardButton(text="📞 Qo'ng'iroq qilingan (kotargan)")],
    [KeyboardButton(text="🔄 Qayta bog'lanildi")],
    [KeyboardButton(text="❌ Qo'ng'iroq qilinmadi")],
    [KeyboardButton(text="🚫 Aktiv emas / noto'g'ri")],
    [KeyboardButton(text="🔙 Ortga")]
], resize_keyboard=True)

admin_kb = ReplyKeyboardMarkup(keyboard=[
    [KeyboardButton(text="📊 Umumiy statistika"), KeyboardButton(text="👥 Operatorlar reytingi")],
    [KeyboardButton(text="➕ Raqam(lar) qo'shish"), KeyboardButton(text="⚙️ Limitni sozlash")],
    [KeyboardButton(text="📉 Kotarmaganlar (Excel)"), KeyboardButton(text="📞 Barcha raqamlar (Excel)")],
    [KeyboardButton(text="👤 Tasdiqlangan operatorlar"), KeyboardButton(text="🆕 Yangi so'rovlar")],
    [KeyboardButton(text="📅 Kunlik Excel"), KeyboardButton(text="📅 Haftalik Excel"), KeyboardButton(text="📅 Oylik Excel")],
    [KeyboardButton(text="🧹 Tozalash"), KeyboardButton(text="🔙 Asosiy Menyu")]
], resize_keyboard=True)

personal_kb = ReplyKeyboardMarkup(keyboard=[
    [KeyboardButton(text="📅 Kunlik hisobot (Excel)")],
    [KeyboardButton(text="📅 Haftalik hisobot (Excel)")],
    [KeyboardButton(text="🔙 Ortga")]
], resize_keyboard=True)

# ================== HANDLERS ==================

# 1. ADMIN PANEL ORTGA QAYTISH (TUZATILDI)
@router.message(F.text == "🔙 Asosiy Menyu")
async def admin_exit_to_main(m: Message, state: FSMContext):
    await state.clear()
    uid = m.from_user.id
    user = db.get_user(uid)
    
    if user and user[4]: # Agar tasdiqlangan user bo'lsa
        is_online = bool(user[5])
        await m.answer("Asosiy menyuga qaytildi.", reply_markup=main_kb(uid, is_online))
    else:
        await m.answer("Siz admin emassiz yoki tizimga kirmagansiz.", reply_markup=ReplyKeyboardRemove())

# 2. GLOBAL ORTGA (USERLAR UCHUN)
@router.message(F.text == "🔙 Ortga")
async def global_cancel(m: Message, state: FSMContext):
    current_state = await state.get_state()
    
    # Agar Admin parol kiritish joyida yoki ichki menyularida bo'lsa
    if current_state and (current_state.startswith("AdminSt") or current_state.startswith("CoAdminSt")):
        await state.clear()
        # Adminni admin panelga emas, startga otamiz (yoki admin panelga qaytarish mumkin)
        # Hozircha User menyusiga qaytaramiz, xavfsizlik uchun
        uid = m.from_user.id
        user = db.get_user(uid)
        is_online = bool(user[5]) if user else False
        await m.answer("Bekor qilindi.", reply_markup=main_kb(uid, is_online))
        return

    await state.clear()
    uid = m.from_user.id
    user = db.get_user(uid)
    
    # User online/offline holatini olamiz
    if user and user[4]: 
        is_online = bool(user[5])
        await m.answer("Asosiy menyu:", reply_markup=main_kb(uid, is_online))
    else: 
        await m.answer("Bekor qilindi.", reply_markup=ReplyKeyboardRemove())

@router.message(CommandStart())
async def cmd_start(m: Message, state: FSMContext):
    await state.clear()
    uid = m.from_user.id
    user = db.get_user(uid)
    
    if user and user[4]: 
        is_online = bool(user[5])
        await m.answer("Xush kelibsiz!", reply_markup=main_kb(uid, is_online))
    elif user and not user[4]: 
        await m.answer("⏳ Hisobingiz admin tasdiqlashini kutmoqda.")
    else:
        kb = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="📱 Raqamni yuborish", request_contact=True)]], resize_keyboard=True)
        await m.answer("Assalomu alaykum! Ro'yxatdan o'tish uchun raqamingizni yuboring:", reply_markup=kb)

@router.message(F.contact)
async def reg_contact(m: Message, state: FSMContext):
    phone = m.contact.phone_number
    if not phone.startswith("+"): phone = "+" + phone
    await state.update_data(phone=phone)
    await m.answer("Ism va familiyangizni kiriting:", reply_markup=back_kb)
    await state.set_state(Reg.name)

@router.message(Reg.name)
async def reg_name(m: Message, state: FSMContext):
    data = await state.get_data()
    if db.add_user(m.from_user.id, m.text, data['phone']):
        await m.answer("✅ So'rov yuborildi. Kuting.", reply_markup=ReplyKeyboardRemove())
    else: 
        await m.answer("Siz allaqachon ro'yxatdasiz.")
    await state.clear()

# --- OPERATOR QISMI ---

@router.message(F.text.startswith("🟢 Ish vaqti:"))
async def toggle_online(m: Message):
    uid = m.from_user.id
    user = db.get_user(uid)
    if not user or not user[4]: return await m.answer("Ruxsat yo'q.")
    
    new_status = not bool(user[5])
    db.set_online(uid, new_status)
    await m.answer("Holat o'zgardi.", reply_markup=main_kb(uid, new_status))

@router.message(F.text == "📞 Nomer olish")
async def get_number(m: Message, state: FSMContext):
    uid = m.from_user.id
    user = db.get_user(uid)
    
    if not user or not user[5]: 
        return await m.answer("Avval Online bo'ling.")
    
    # Tugatilmagan raqam tekshiruvi
    if user[6]: 
        active_details = db.get_number_details(user[6])
        if active_details:
            ph, nm, tg, ex = active_details
            msg = (
                f"⚠️ <b>Sizda tugatilmagan raqam bor!</b>\n"
                f"Avval shu raqamni yakunlang.\n\n"
                f"📞 <b>Mijoz:</b>\n"
                f"👤 <b>Ism:</b> {nm if nm else 'Topilmadi'}\n"
                f"📱 <b>Tel:</b> <code>{ph}</code>\n"
                f"✈️ <b>Tg:</b> {tg if tg else 'Yoq'}\n"
                f"📝 <b>Info:</b> {ex if ex else 'Yoq'}"
            )
            await m.answer(msg, reply_markup=call_action_kb)
            await state.set_state(Call.waiting_for_action)
            return
        else:
            db.set_current_number(uid, None)
    
    # Limit
    limit = db.get_limit()
    count = db.get_today_count(uid)
    if count >= limit: 
        return await m.answer(f"⛔️ Kunlik limit tugadi ({count}/{limit}).")
    
    # Yangi raqam
    data = db.get_free_number_full(uid) 
    if data:
        phone, name, tg, extra = data
        db.set_current_number(uid, phone)
        msg = (
            f"📞 <b>Yangi mijoz:</b>\n\n"
            f"👤 <b>Ism:</b> {name if name else 'Topilmadi'}\n"
            f"📱 <b>Tel:</b> <code>{phone}</code>\n"
            f"✈️ <b>Tg:</b> {tg if tg else 'Yoq'}\n"
            f"📝 <b>Info:</b> {extra if extra else 'Yoq'}\n\n"
            f"<i>Qo'ng'iroq qilib, natijani tanlang:</i>"
        )
        await m.answer(msg, reply_markup=call_action_kb)
        await state.set_state(Call.waiting_for_action)
    else: 
        await m.answer("⚠️ Bazada bo'sh raqam qolmadi.")

@router.message(F.text == "✍️ Raqam kiritish")
async def manual_enter_start(m: Message, state: FSMContext):
    user = db.get_user(m.from_user.id)
    if not user or not user[5]: return await m.answer("Avval Online bo'ling.")
    
    await m.answer("Mijoz raqamini kiriting:", reply_markup=back_kb)
    await state.set_state(ManualEntry.phone)

@router.message(ManualEntry.phone)
async def manual_enter_save(m: Message, state: FSMContext):
    phone = m.text.strip()
    if len(phone) < 7: return await m.answer("Noto'g'ri raqam.", reply_markup=back_kb)
    
    await state.update_data(phone=phone)
    await m.answer(f"Raqam: <b>{phone}</b>\nNatijani tanlang:", reply_markup=call_action_kb)
    await state.set_state(ManualEntry.action)

# --- NATIJA QABUL QILISH ---
@router.message(Call.waiting_for_action)
@router.message(ManualEntry.action)
async def log_call_result(m: Message, state: FSMContext):
    st_map = {
        "📞 Qo'ng'iroq qilingan (kotargan)": "success", 
        "🔄 Qayta bog'lanildi": "recalled", 
        "❌ Qo'ng'iroq qilinmadi": "no_answer", 
        "🚫 Aktiv emas / noto'g'ri": "invalid"
    }
    
    if m.text not in st_map: return await m.answer("Iltimos, tugmalardan birini tanlang.")

    uid = m.from_user.id
    current_state = await state.get_state()
    
    phone = None
    if str(current_state) == "ManualEntry:action":
        d = await state.get_data()
        phone = d.get('phone')
    else:
        u = db.get_user(uid)
        if u and u[6]: phone = u[6]
        else: return await m.answer("Faol raqam topilmadi.")

    status = st_map[m.text]
    await state.update_data(phone=phone, status=status)

    if status in ["success", "recalled"]:
        await m.answer("Mijoz ismi:", reply_markup=back_kb)
        await state.set_state(Call.name)
    else:
        db.add_call({'op_id': uid, 'phone': phone, 'status': status})
        if str(current_state) != "ManualEntry:action": 
            db.set_current_number(uid, None)
        await state.clear()
        await m.answer("✅ Natija saqlandi!", reply_markup=main_kb(uid, True))

# --- ANKETA ---
@router.message(Call.name)
async def c_name(m: Message, state: FSMContext): 
    await state.update_data(name=m.text)
    await m.answer("Yosh:", reply_markup=back_kb)
    await state.set_state(Call.age)

@router.message(Call.age)
async def c_age(m: Message, state: FSMContext): 
    val = int(m.text) if m.text.isdigit() else 0
    await state.update_data(age=val)
    await m.answer("Bo'y:", reply_markup=back_kb)
    await state.set_state(Call.height)

@router.message(Call.height)
async def c_height(m: Message, state: FSMContext): 
    val = int(m.text) if m.text.isdigit() else 0
    await state.update_data(height=val)
    await m.answer("Vazn:", reply_markup=back_kb)
    await state.set_state(Call.weight)

@router.message(Call.weight)
async def c_weight(m: Message, state: FSMContext):
    val = int(m.text) if m.text.isdigit() else 0
    await state.update_data(weight=val)
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Ha (Qiziqdi)", callback_data="int_ha"), 
        InlineKeyboardButton(text="❌ Yo'q", callback_data="int_yoq")
    ]])
    await m.answer("Mijoz qiziqdimi?", reply_markup=kb)
    await state.set_state(Call.interest)

@router.callback_query(Call.interest)
async def c_interest(c: CallbackQuery, state: FSMContext):
    ans = "Ha" if c.data == "int_ha" else "Yo'q"
    d = await state.get_data()
    uid = c.from_user.id
    
    db.add_call({
        'op_id': uid, 'phone': d['phone'], 'status': d['status'], 
        'name': d.get('name'), 'age': d.get('age'), 
        'height': d.get('height'), 'weight': d.get('weight'), 'interest': ans
    })
    
    user = db.get_user(uid)
    if user and user[6] == d['phone']: 
        db.set_current_number(uid, None)
        
    await state.clear()
    await c.message.edit_text(f"Qiziqish: {ans}")
    await c.message.answer("✅ Saqlandi!", reply_markup=main_kb(uid, True))

# --- QO'SHIMCHA USER FUNKSIYALARI ---
@router.message(F.text == "📉 Bugungi qabul qilinmagan")
async def today_no(m: Message):
    nums = db.get_today_no_answers(m.from_user.id)
    if not nums:
        await m.answer("Bugun barcha raqamlarga javob berilgan.")
    else:
        lines = []
        for idx, r in enumerate(nums, 1): lines.append(f"{idx}. {r[0]}")
        await m.answer("📉 <b>Bugun javob bermagan raqamlar:</b>\n" + "\n".join(lines))

@router.message(F.text == "📈 Shaxsiy statistikam")
async def my_st(m: Message): 
    await m.answer("Davrni tanlang:", reply_markup=personal_kb)

@router.message(F.text.in_(["📅 Kunlik hisobot (Excel)", "📅 Haftalik hisobot (Excel)"]))
async def my_ex(m: Message): 
    p = "kunlik" if "Kunlik" in m.text else "haftalik"
    await generate_excel(m, p, m.from_user.id)

# ================== ADMIN QISMI ==================

@router.message(Command("admin"))
async def admin_ent(m: Message, state: FSMContext):
    if m.from_user.id in ADMIN_IDS: 
        await m.answer("Parol:", reply_markup=back_kb)
        await state.set_state(AdminSt.password)

@router.message(AdminSt.password)
async def admin_chk(m: Message, state: FSMContext):
    if m.text == ADMIN_PASSWORD: 
        await m.answer("Admin Panel", reply_markup=admin_kb)
        await state.clear()
    else: 
        await m.answer("Xato.")

@router.message(F.text == "➕ Raqam(lar) qo'shish")
async def adm_add(m: Message, state: FSMContext):
    if m.from_user.id not in ADMIN_IDS: return
    await m.answer("Excel fayl (.xlsx) yuboring:", reply_markup=back_kb)
    await state.set_state(AdminSt.add_nums)

@router.message(AdminSt.add_nums, F.document)
async def adm_upload_excel(m: Message, state: FSMContext, bot: Bot):
    if not m.document.file_name.endswith('.xlsx'):
        return await m.answer("Faqat .xlsx fayl yuboring!")
    
    file_id = m.document.file_id
    file = await bot.get_file(file_id)
    destination = f"import_{m.from_user.id}.xlsx"
    await bot.download_file(file.file_path, destination)
    
    added, updated = 0, 0 
    try:
        wb = load_workbook(destination, data_only=True)
        ws = wb.active
        
        anchor_text = "ismingizni_yozing"
        start_col = 0
        start_row = 0
        found = False
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            for c_idx, cell_value in enumerate(row):
                if cell_value and anchor_text in str(cell_value).lower():
                    start_row = r_idx; start_col = c_idx; found = True; break
            if found: break
            
        if not found: await m.answer("⚠️ 'ismingizni_yozing' topilmadi!"); return

        for row in ws.iter_rows(min_row=start_row + 1, values_only=True):
            if not row: continue
            def get_cell(offset):
                idx = start_col + offset
                return str(row[idx]).strip() if idx < len(row) and row[idx] else ""

            name = get_cell(0)
            raw_phone = get_cell(1)
            tg = get_cell(2)
            info = get_cell(3)
            extra_phone = get_cell(4)
            full_extra_info = f"{info} | 2-tel: {extra_phone}" if extra_phone else info
            phone = re.sub(r'[^\d+]', '', raw_phone)
            
            if len(phone) < 7: continue
            res = db.add_full_number(phone, name, tg, full_extra_info)
            if res == True: added += 1
            elif res == "updated": updated += 1
            
        await m.answer(f"✅ Yuklandi!\n🆕 Qo'shildi: {added}\n♻️ Yangilandi: {updated}", reply_markup=admin_kb)
    except Exception as e: await m.answer(f"Xatolik: {e}")
    finally: os.remove(destination)
    await state.clear()

async def generate_excel(m, period, op_id):
    now = datetime.now()
    if period == "kunlik": s = now.replace(hour=0, minute=0, second=0); e = now + timedelta(days=1)
    elif period == "haftalik": s = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0); e = s + timedelta(days=7)
    elif period == "oylik": s = now.replace(day=1, hour=0, minute=0, second=0); e = (s.replace(month=s.month+1) if s.month < 12 else s.replace(year=s.year+1, month=1))
    
    rows = db.get_calls_stats(s, e, op_id)
    if not rows: return await m.answer("Ma'lumot yo'q.")
    wb = Workbook(); ws = wb.active
    ws.append(["Op ID", "Op", "Tel", "Status", "Mijoz", "Yosh", "Bo'y", "Vazn", "Qiziqish", "Vaqt"])
    for r in rows: ws.append(list(r))
    fn = f"Rep_{period}_{m.from_user.id}.xlsx"
    wb.save(fn); await m.answer_document(FSInputFile(fn)); os.remove(fn)

@router.message(F.text.in_(["📅 Kunlik Excel", "📅 Haftalik Excel", "📅 Oylik Excel"]))
async def adm_exc(m: Message): 
    if m.from_user.id in ADMIN_IDS: 
        p = "kunlik" if "Kunlik" in m.text else "haftalik" if "Haftalik" in m.text else "oylik"
        await generate_excel(m, p, None)

@router.message(F.text == "📉 Kotarmaganlar (Excel)")
async def adm_noans(m: Message):
    if m.from_user.id not in ADMIN_IDS: return
    rows = db.get_no_answer_numbers()
    wb = Workbook(); ws = wb.active; ws.append(["Tel"])
    for r in rows: ws.append(list(r))
    fn = "NoAns.xlsx"; wb.save(fn); await m.answer_document(FSInputFile(fn)); os.remove(fn)

@router.message(F.text == "📞 Barcha raqamlar (Excel)")
async def adm_allnums(m: Message):
    if m.from_user.id not in ADMIN_IDS: return
    data = db.get_all_numbers()
    wb = Workbook(); ws = wb.active; ws.append(["Tel", "Ishlatilgan", "Operator", "Vaqt"])
    for r in data: ws.append(list(r))
    fn = "AllNums.xlsx"; wb.save(fn); await m.answer_document(FSInputFile(fn)); os.remove(fn)

@router.message(F.text == "📊 Umumiy statistika")
async def adm_stats(m: Message):
    if m.from_user.id not in ADMIN_IDS: return
    t, s, n, i = db.get_general_stats()
    msg = f"<b>📊 Statistika</b>\n📞 Jami: {t}\n✅ Kotargan: {s}\n❌ Yo'q: {n}\n🚫 Yaroqsiz: {i}"
    await m.answer(msg)

@router.message(F.text == "👥 Operatorlar reytingi")
async def adm_rank(m: Message):
    if m.from_user.id not in ADMIN_IDS: return
    lines = []
    for r in db.get_operator_ranking(): lines.append(f"👤 {r[0]}: <b>{r[3]}</b>")
    await m.answer("<b>🏆 Reyting:</b>\n" + "\n".join(lines))

@router.message(F.text == "⚙️ Limitni sozlash")
async def set_limit_start(m: Message, state: FSMContext):
    if m.from_user.id not in ADMIN_IDS: return
    await m.answer(f"Limit: {db.get_limit()}", reply_markup=back_kb); await state.set_state(AdminSt.set_limit)

@router.message(AdminSt.set_limit)
async def set_limit_save(m: Message, state: FSMContext):
    if m.text.isdigit(): db.set_limit(int(m.text)); await m.answer("OK", reply_markup=admin_kb); await state.clear()

@router.message(F.text == "🆕 Yangi so'rovlar")
async def adm_new(m: Message):
    pending = [u for u in db.get_all_users() if not u[4]]
    if not pending: return await m.answer("Hozircha yangi so'rovlar yo'q.")
    for u in pending:
        kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="✅", callback_data=f"app_{u[1]}"), InlineKeyboardButton(text="❌", callback_data=f"rej_{u[1]}")]])
        await m.answer(f"👤 {u[2]}\n📱 {u[3]}", reply_markup=kb)

@router.callback_query(F.data.startswith(("app_", "rej_", "del_")))
async def adm_act(c: CallbackQuery):
    act, tid = c.data.split("_"); tid = int(tid)
    if act == "app": 
        db.approve_user(tid, True); await bot.send_message(tid, "✅ Hisobingiz tasdiqlandi!"); await c.message.edit_text("✅ OK")
    elif act == "rej": 
        db.delete_user(tid); await bot.send_message(tid, "❌ Rad etildi."); await c.message.edit_text("❌ DEL")
    elif act == "del": 
        db.delete_user(tid); await c.message.edit_text("🗑 Deleted")

@router.message(F.text == "👤 Tasdiqlangan operatorlar")
async def adm_appr(m: Message):
    if m.from_user.id not in ADMIN_IDS: return
    for u in [x for x in db.get_all_users() if x[4]]:
        kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"del_{u[1]}")]])
        await m.answer(f"{'🟢' if u[5] else '🔴'} {u[2]} | {u[3]}", reply_markup=kb)

@router.message(F.text == "🧹 Tozalash")
async def adm_clear_ask(m: Message):
    if m.from_user.id in ADMIN_IDS: 
        kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="✅ Ha", callback_data="clear_yes"), InlineKeyboardButton(text="❌ Yo'q", callback_data="clear_no")]])
        await m.answer("Tozalansinmi?", reply_markup=kb)

@router.callback_query(F.data.startswith("clear_"))
async def adm_clear_confirm(c: CallbackQuery):
    if "yes" in c.data: db.clear_numbers_and_calls(); await c.message.edit_text("Baza tozalandi.")
    else: await c.message.edit_text("Bekor qilindi.")

# --- CO-ADMIN & SEARCH ---
@router.message(Command("coadmin"))
async def coadmin_ent(m: Message, state: FSMContext): 
    await m.answer("Parol:", reply_markup=back_kb); await state.set_state(CoAdminSt.password)

@router.message(CoAdminSt.password)
async def coadmin_chk(m: Message, state: FSMContext):
    if m.text == COADMIN_PASSWORD: 
        db.make_coadmin(m.from_user.id, True); await m.answer("✅ Siz Co-Adminsiz!", reply_markup=main_kb(m.from_user.id, True)); await state.clear()
    else: await m.answer("Xato parol.")

@router.message(F.text == "🔍 Raqam tekshirish")
async def search_start(m: Message, state: FSMContext): 
    await m.answer("Oxirgi 4 ta raqam:", reply_markup=back_kb); await state.set_state(Search.query)

@router.message(Search.query)
async def search_process(m: Message, state: FSMContext):
    res = db.search_phone_by_digits(m.text)
    sid = m.from_user.id
    s_user = db.get_user(sid)
    is_priv = (sid in ADMIN_IDS) or (s_user and len(s_user)>7 and s_user[7])
    
    if not res: await m.answer("Topilmadi.", reply_markup=main_kb(sid, True))
    else:
        for r in res[:5]:
            ph, op, dt, st, cn, age, h, w, intr, op_id = r
            txt = f"📱 {ph}\n👤 Op: {op}\n📊 Stat: {st}\n🕒 Vaqt: {dt}"
            if is_priv and op_id and int(op_id) != sid:
                kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="✅ Eslatma yuborish", callback_data=f"ntf_y_{op_id}_{ph}"), InlineKeyboardButton(text="❌ Yopish", callback_data="ntf_n")]])
                await m.answer(txt, reply_markup=kb)
            else: await m.answer(txt)
        await m.answer("Qidiruv yakunlandi.", reply_markup=main_kb(sid, True))
    await state.clear()

# --- YANGILANGAN ESLATMA TIZIMI (MUKAMMAL) ---
@router.callback_query(F.data.startswith("ntf_"))
async def process_notification_callback(c: CallbackQuery):
    if c.data == "ntf_n":
        await c.message.delete()
        return

    # Eslatma yuborish logikasi
    try:
        parts = c.data.split("_")
        op_id = int(parts[2])
        phone = parts[3]
        
        # Operatorga xabar yuborish
        await bot.send_message(
            chat_id=op_id, 
            text=f"🔔 <b>DIQQAT, ESLATMA!</b>\n\nSiz ishlagan ushbu mijoz: <b>{phone}</b>\nAdmin tomonidan qayta tekshirildi.\nIltimos, qayta bog'laning!"
        )
        
        # Admin xabarini o'zgartirish (Confirmation)
        await c.message.edit_text(f"{c.message.text}\n\n✅ <b>Eslatma muvaffaqiyatli yuborildi!</b>", reply_markup=None)
        await c.answer("Yuborildi!", show_alert=True)
        
    except Exception as e:
        await c.answer(f"Xatolik: Operator botni bloklagan bo'lishi mumkin.\n({e})", show_alert=True)

async def main():
    print("Bot Aiogram 3 da ishga tushdi...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
